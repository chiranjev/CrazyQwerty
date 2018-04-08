# -*- coding: utf-8 -*-
"""
Created on Wed Apr  4 10:34:44 2018

@author: Chiranjev Koul
"""

import openpyxl

wb = openpyxl.load_workbook('dict.xlsx')

sheet_names = wb.get_sheet_names()

sheet = wb.get_sheet_by_name(sheet_names[0])
print(sheet)

dict = {}

word = []
meaning = []

for i in range(1,601,2):
    dict[sheet.cell(row=i,column=1).value] = sheet.cell(row=i+1,column=1).value
    word.append(sheet.cell(row=i,column=1).value)
    meaning.append(sheet.cell(row=i+1,column=1).value)
     
print(dict)
print(len(dict))

print(word)
print(len(word))
print(meaning)
print(len(meaning))